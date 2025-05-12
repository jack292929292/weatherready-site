import pandas as pd
import numpy as np
import requests
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Step 1: Build BOM URL for current month's Perth Metro CSV
now = datetime.now()
year_month = now.strftime("%Y%m")
url = f"https://www.bom.gov.au/climate/dwo/{year_month}/text/IDCJDW6111.{year_month}.csv"

# Step 2: Fetch the CSV with proper headers to spoof browser
headers = {
    "User-Agent": "Mozilla/5.0"
}
response = requests.get(url, headers=headers)
if response.status_code != 200:
    print(f"❌ Failed to fetch BOM data. Status code: {response.status_code}")
    exit()

# Step 3: Load the CSV into a DataFrame
from io import StringIO
data = pd.read_csv(StringIO(response.text), encoding='utf-8', sep=",", skiprows=6)

# Step 4: Find most recent valid date with a max temp
data["Date"] = pd.to_datetime(data["Date"], errors='coerce')
data = data.dropna(subset=["Date"])
data["Maximum temperature (°C)"] = pd.to_numeric(data["Maximum temperature (°C)"], errors='coerce')
recent_row = data.dropna(subset=["Maximum temperature (°C)"]).iloc[-1]

obs_date = recent_row["Date"].date()
max_temp = recent_row["Maximum temperature (°C)"]

# Step 5: Load Excel and update WeatherImport sheet
file_path = "WeatherReady2025_POWERQUERY_READY.xlsx"
book = load_workbook(file_path)
ws = book["WeatherImport"]
ws["A2"] = obs_date.strftime("%Y-%m-%d")
ws["B2"] = max_temp

# Step 6: Recalculate MaxPredict2 in Sheet2
sheet2_df = pd.read_excel(file_path, sheet_name="Sheet2")
sheet2_df["YearDay"] = pd.to_datetime(sheet2_df["Date"]).dt.dayofyear

grouped = sheet2_df.groupby("YearDay")
weighted_avg = grouped["MaxWeightedAverage"].mean()
weighted_std = grouped["MaxWeightedAverage"].std()
ndmd_avg = grouped["NDMDAverage"].mean()
ndmd_std = grouped["NDMDAverage"].std()

maxpredict2 = []
for idx, row in sheet2_df.iterrows():
    yd = row["YearDay"]
    if pd.notna(yd) and yd in weighted_avg and yd in ndmd_avg:
        try:
            mean = (
                (weighted_avg[yd] / (weighted_std[yd]**2) + (row["MaxPredict1"] + ndmd_avg[yd]) / (ndmd_std[yd]**2)) /
                ((1 / (weighted_std[yd]**2)) + (1 / (ndmd_std[yd]**2)))
            )
            std_dev = np.sqrt(1 / ((1 / (weighted_std[yd]**2)) + (1 / (ndmd_std[yd]**2))))
            maxpredict2.append(np.random.normal(loc=mean, scale=std_dev))
        except:
            maxpredict2.append(None)
    else:
        maxpredict2.append(None)

sheet2_df["MaxPredict2"] = maxpredict2

# Step 7: Save updated Sheet2 and Excel file
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    writer.book = book
    sheet2_df.to_excel(writer, sheet_name="Sheet2", index=False)

book.save(file_path)

print(f"✅ Refreshed for {obs_date} — Max Temp: {max_temp}°C")
